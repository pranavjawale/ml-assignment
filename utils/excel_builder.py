"""
Excel and mapping file writer.

Produces:
  <tag>_speaker-name-analysis.xlsx
      SHEET1 — canonical speakers (stats)
      SHEET2 — per-PDF speaker x file matrix
      SHEET3 — dedup review (canonical groups / wildcards / roles)
  <tag>_spk-id-mapping.txt  — tab-separated raw-name -> spk_id mapping
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .extractor import PdfExtractor
from .deduplicator import (
    SpeakerDeduplicator, speaker_type, extract_title_and_name, strip_titles,
)


# ── Style helpers ─────────────────────────────────────────────────────────────

_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT     = Alignment(horizontal="left",   vertical="center")
_THIN     = Side(style="thin", color="AAAAAA")
_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header(cell, align=None) -> None:
    cell.font      = _HDR_FONT
    cell.fill      = _HDR_FILL
    cell.border    = _BORDER
    cell.alignment = align or _CENTER


def _style_cell(cell, fill=None, align=None) -> None:
    cell.border    = _BORDER
    cell.alignment = align or _CENTER
    if fill:
        cell.fill  = fill


def _row_fill(row_idx: int):
    return _ALT_FILL if row_idx % 2 == 0 else None


# ── Public API ────────────────────────────────────────────────────────────────

def write_spk_id_mapping(
    deduplicator: SpeakerDeduplicator,
    output_path: str,
) -> None:
    """Write tab-separated raw-name -> spk_id mapping file."""
    raw_spk_id = deduplicator.raw_spk_id
    with open(output_path, "w", encoding="utf-8") as f:
        for raw_name in sorted(raw_spk_id, key=lambda n: (raw_spk_id[n], n)):
            f.write(f"{raw_name}\t{raw_spk_id[raw_name]}\n")
    print(f"Written: {output_path}  ({len(raw_spk_id)} entries)")


def build_speaker_excel(
    extractor: PdfExtractor,
    deduplicator: SpeakerDeduplicator,
    output_path: str,
) -> None:
    """
    Build the speaker-name-analysis workbook with three sheets.

    SHEET1 columns:
      Speaker Name | Speaker ID | Title/Honorifics | First Name | Middle Name
      | Surname | Type | Total Words | File Indices | # Files
      | Shortest Segment | Longest Segment

    SHEET2:
      PDF Index | Total Speakers | <one column per canonical speaker>

    SHEET3 (dedup review — same information as the old deduplicate_candidates.txt):
      Section | Group | Member Type | Speaker Name | Speaker ID | Total Words
      | # Files | File Indices | Stripped Name | Notes | Reviewer Note
    """
    canonical = deduplicator.canonical_speakers
    wb = openpyxl.Workbook()

    _build_sheet1(wb, deduplicator, canonical)
    _build_sheet2(wb, extractor, deduplicator, canonical)
    _build_sheet3(wb, extractor, deduplicator)

    wb.save(output_path)
    print(f"Written: {output_path}  "
          f"(SHEET1: {len(canonical)} rows, "
          f"SHEET2: {len(extractor.raw_data)} rows x {len(canonical)} speaker cols, "
          f"SHEET3: dedup review)")


# ── Internal builders ─────────────────────────────────────────────────────────

def _build_sheet1(
    wb: openpyxl.Workbook,
    dedup: SpeakerDeduplicator,
    canonical: list[str],
) -> None:
    ws = wb.active
    ws.title = "SHEET1"

    headers = [
        "Speaker Name", "Speaker ID",
        "Title / Honorifics", "First Name", "Middle Name", "Surname",
        "Type", "Total Words Spoken", "File Indices", "Number of Files",
        "Shortest Segment (words)", "Longest Segment (words)",
    ]
    col_widths = [55, 14, 30, 18, 18, 22, 16, 20, 40, 16, 22, 22]

    ws.row_dimensions[1].height = 30
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _style_header(ws.cell(1, col, h))
        ws.column_dimensions[get_column_letter(col)].width = w

    for row_idx, canon in enumerate(canonical, start=2):
        fill         = _row_fill(row_idx)
        file_indices = sorted(dedup.canon_files.get(canon, set()))
        title, first, middles, surname = extract_title_and_name(canon)

        _style_cell(ws.cell(row_idx, 1,  canon),                                      fill, _LEFT)
        _style_cell(ws.cell(row_idx, 2,  dedup.canon_spk_id.get(canon, "")),          fill)
        _style_cell(ws.cell(row_idx, 3,  title),                                      fill, _LEFT)
        _style_cell(ws.cell(row_idx, 4,  first),                                      fill)
        _style_cell(ws.cell(row_idx, 5,  middles),                                    fill)
        _style_cell(ws.cell(row_idx, 6,  surname),                                    fill)
        _style_cell(ws.cell(row_idx, 7,  speaker_type(canon)),                        fill)
        _style_cell(ws.cell(row_idx, 8,  dedup.canon_words.get(canon, 0)),            fill)
        _style_cell(ws.cell(row_idx, 9,  ", ".join(str(i) for i in file_indices)),    fill, _LEFT)
        _style_cell(ws.cell(row_idx, 10, len(file_indices)),                          fill)
        _style_cell(ws.cell(row_idx, 11, dedup.canon_seg_min.get(canon, "")),         fill)
        _style_cell(ws.cell(row_idx, 12, dedup.canon_seg_max.get(canon, "")),         fill)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(canonical) + 1}"


def _build_sheet2(
    wb: openpyxl.Workbook,
    ext: PdfExtractor,
    dedup: SpeakerDeduplicator,
    canonical: list[str],
) -> None:
    ws = wb.create_sheet("SHEET2")

    fixed_headers = ["PDF Index", "Total Speakers in File"]
    all_headers   = fixed_headers + canonical

    ws.row_dimensions[1].height = 80
    for col, h in enumerate(all_headers, start=1):
        c = ws.cell(1, col, h)
        _style_header(c)
        if col > 2:
            c.alignment = Alignment(
                horizontal="center", vertical="bottom",
                text_rotation=90, wrap_text=False,
            )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    for col in range(3, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8

    for row_idx, pdf_idx in enumerate(sorted(ext.raw_data.keys()), start=2):
        per_file = dedup.canon_per_pdf.get(pdf_idx, {})
        fill     = _row_fill(row_idx)
        _style_cell(ws.cell(row_idx, 1, pdf_idx),       fill)
        _style_cell(ws.cell(row_idx, 2, len(per_file)), fill)
        for col, canon in enumerate(canonical, start=3):
            wc = per_file.get(canon)
            _style_cell(ws.cell(row_idx, col, wc if wc else None), fill)

    ws.freeze_panes = "C2"


def _build_sheet3(
    wb: openpyxl.Workbook,
    ext: PdfExtractor,
    dedup: SpeakerDeduplicator,
) -> None:
    """
    Dedup review sheet — the same information the old deduplicate_candidates.txt
    carried, laid out as a filterable table:

      Section A : canonical speaker groups (multi-variant + singletons)
      Section B : unresolved wildcards (surname-only, ambiguous group match)
      Section C : role / generic labels

    A blank "Reviewer Note" column is provided for human corrections.
    """
    ws = wb.create_sheet("SHEET3")

    gwt       = ext.global_word_totals
    file_sets = ext.global_file_sets
    groups    = dedup.groups
    wildcards = dedup.wildcards
    roles     = dedup.roles

    multi  = [g for g in groups if len(g) > 1]
    single = [g for g in groups if len(g) == 1]

    headers = [
        "Section", "Group", "Member Type", "Speaker Name", "Speaker ID",
        "Total Words", "Number of Files", "File Indices", "Stripped Name",
        "Notes", "Reviewer Note",
    ]
    col_widths = [16, 8, 22, 50, 12, 14, 16, 26, 34, 44, 28]
    ws.row_dimensions[1].height = 30
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _style_header(ws.cell(1, col, h))
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 2

    def emit(section: str, group: str, mtype: str, name: str, notes: str = "") -> None:
        nonlocal row
        fill  = _row_fill(row)
        files = sorted(file_sets.get(name, set()))
        _style_cell(ws.cell(row, 1,  section),                             fill, _LEFT)
        _style_cell(ws.cell(row, 2,  group),                               fill)
        _style_cell(ws.cell(row, 3,  mtype),                               fill, _LEFT)
        _style_cell(ws.cell(row, 4,  name),                                fill, _LEFT)
        _style_cell(ws.cell(row, 5,  dedup.raw_spk_id.get(name, "")),      fill)
        _style_cell(ws.cell(row, 6,  gwt.get(name, 0)),                    fill)
        _style_cell(ws.cell(row, 7,  len(files)),                          fill)
        _style_cell(ws.cell(row, 8,  ", ".join(str(f) for f in files)),    fill, _LEFT)
        _style_cell(ws.cell(row, 9,  strip_titles(name)),                  fill, _LEFT)
        _style_cell(ws.cell(row, 10, notes),                               fill, _LEFT)
        _style_cell(ws.cell(row, 11, ""),                                  fill, _LEFT)
        row += 1

    # Section A — multi-variant groups (canonical first, then variants)
    for idx, g in enumerate(multi, 1):
        gid       = f"G{idx:03d}"
        total     = sum(gwt.get(n, 0) for n in g)
        canonical = g[0]
        for n in g:
            if n == canonical:
                emit("A: Group", gid, "CANONICAL", n,
                     f"{len(g)} variants, {total:,} words total")
            else:
                emit("A: Group", gid, "variant", n)

    # Section A — singletons (groups with no variants found)
    for g in single:
        emit("A: Singleton", "", "singleton", g[0])

    # Section B — unresolved wildcards (surname-only, multiple group matches).
    # Each candidate lists the file IDs where that speaker appears, so the
    # reviewer can compare against the wildcard's own File Indices column.
    for wname, hits in sorted(wildcards, key=lambda x: -gwt.get(x[0], 0)):
        if hits:
            parts = []
            for h in hits:
                group_files = sorted(set().union(*(file_sets.get(n, set()) for n in h)))
                fids = ", ".join(str(f) for f in group_files)
                parts.append(f"{strip_titles(h[0])} (files: {fids})")
            matches = "; ".join(parts)
        else:
            matches = "no group matched"
        emit("B: Wildcard", "", "wildcard (unresolved)", wname,
             f"possible matches: {matches}")

    # Section C — role / generic labels
    for n in sorted(roles, key=lambda n: -gwt.get(n, 0)):
        emit("C: Role", "", "role", n)

    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
