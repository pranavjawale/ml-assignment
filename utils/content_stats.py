"""
Transcript content statistics — spaCy-based analysis of raw transcripts.

Entry point:
    run_content_stats(transcript_dir, output_dir, tag)

Produces <tag>_transcript_content_analysis.xlsx with five sheets:
  Sheet 1 — Overview          (speaker count + sentence complete/partial stats)
  Sheet 2 — Category Counts   (sentences containing numerals, dates, times, years, units)
  Sheet 3 — Character Freq    (every character in spoken text, ranked by count)
  Sheet 4 — Word Frequency    (lowercased tokens, ranked by count)
  Sheet 5 — NSW Frequency     (Acronyms / LSEQ / Abbreviations, ranked by count)
"""

import os
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Regex patterns ────────────────────────────────────────────────────────────

# Matches a speaker-label line produced by --initial-transcript-extraction
_SPEAKER_RE  = re.compile(r"^([A-Z][A-Z0-9 .\-']+):\s*$")
_ELLIPSIS_RE = re.compile(r"\.\.\.\s*$")

# For word frequency: strip leading/trailing non-letter/apostrophe chars
_STRIP_PUNCT = re.compile(r"^[^a-zA-Z''’]+|[^a-zA-Z''’]+$")

# Sentence content categories (for Sheet 2 — used with .search() on whole sentences).
# _TIME_RE and _DATE_RE are defined below in the NSW patterns section and reused here.
_NUMERAL_RE = re.compile(r"\d")
_YEAR_RE    = re.compile(r"\b(1[7-9]\d{2}|20\d{2})\b")
_MEASURE_RE = re.compile(
    r"\b\d[\d,.]*\s*"
    r"(?:km|kms|kilomet(?:re|er)s?|mile|miles|met(?:re|er)s?|"
    r"feet|foot|ft|yard|yards|cm|mm|inch|inches|"
    r"kg|kgs|gram|grams|tonne|tonnes|litre|litres|liter|liters|"
    r"sq\.?\s*(?:km|m|ft)|hectare|hectares|acre|acres)\b",
    re.IGNORECASE,
)

# ── NSW token patterns ────────────────────────────────────────────────────────
# These are matched by a single priority-ordered scanner (see _scan_nsw): each
# pattern claims its character span before lower-priority ones, so e.g. "50%" is
# Percentage (not Cardinal "50"), "12/05/2023" is Date, "Rs.5,000" is Money, etc.

_MONTHS = (r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?"
           r"|Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?"
           r"|Nov(?:ember)?|Dec(?:ember)?)")

# Bracketed: any [ ... ] phrase — transcription annotations such as [UNCLEAR],
# [inaudible], [NO AUDIO] that are removed (→ empty) in V3.
_BRACKET_RE = re.compile(r"\[[^\]]*\]")
# Email: local@domain.tld
_EMAIL_RE = re.compile(r"\b[\w.+-]+@[\w-]+\.[\w.-]+\b")
# Hashtag: #word
_HASHTAG_RE = re.compile(r"#\w+")
# Money: currency symbol/code + amount (optionally crore/lakh/etc.), or amount + currency word
_MONEY_RE = re.compile(
    r"(?:Rs\.?|₹|INR|US\$|\$|USD|£|€)\s?\d[\d,]*(?:\.\d+)?"
    r"(?:\s?(?:crores?|lakhs?|millions?|billions?|thousand))?"
    r"|\b\d[\d,]*(?:\.\d+)?\s?(?:rupees?|dollars?|crores?|lakhs?|pounds?|euros?)\b",
    re.IGNORECASE,
)
# Percentage: 50%, 12.5 %, 50 per cent
_PERCENT_RE = re.compile(r"\b\d[\d,]*(?:\.\d+)?\s?(?:%|per\s?cent)", re.IGNORECASE)
# Telephone: +91 numbers, separated digit groups, or bare 10-digit Indian mobile
_PHONE_RE = re.compile(
    r"\+\d{1,3}[\s-]?\d[\d\s-]{6,}\d"
    r"|\b\d{3,5}[\s-]\d{5,8}\b"
    r"|\b[6-9]\d{9}\b"
)
# Time: 10:30, 10:30 a.m., 10 am, 10 o'clock
_TIME_RE = re.compile(
    r"\b\d{1,2}:\d{2}(?:\s?[ap]\.?m\.?)?"
    r"|\b\d{1,2}\s?[ap]\.?m\.?"
    r"|\b\d{1,2}\s?o[''`’]?clock",
    re.IGNORECASE,
)
# Date: numeric (12/05/2023, 2023-05-12) or month-name forms (5th August 2023, Aug 5, 2023)
_DATE_RE = re.compile(
    r"\b\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4}\b"
    r"|\b\d{4}-\d{2}-\d{2}\b"
    r"|\b\d{1,2}(?:st|nd|rd|th)?\s+" + _MONTHS + r"\.?(?:,?\s+\d{2,4})?\b"
    r"|\b" + _MONTHS + r"\.?\s+\d{1,2}(?:st|nd|rd|th)?(?:,?\s+\d{2,4})?\b",
    re.IGNORECASE,
)
# Number range: 25-30, 25 – 30, 25 to 30
_NRANGE_RE = re.compile(r"\b\d[\d,]*\s?(?:[-–—]|to)\s?\d[\d,]*\b", re.IGNORECASE)
# Ordinal: 1st, 2nd, 23rd, 4th
_ORDINAL_RE = re.compile(r"\b\d+(?:st|nd|rd|th)\b", re.IGNORECASE)
# LSEQ: dotted initials — A.K., D.Y., B.R.A.
_LSEQ_RE = re.compile(r"\b[A-Z](?:\.[A-Z])+\.?")
# SPLT: alphanumeric tokens that join letters and digits — e.g. "ITV3", "39B", "COVID19".
# The two lookaheads require the whole token to contain at least one letter AND one digit.
_SPLT_RE = re.compile(r"\b(?=[A-Za-z\d]*[A-Za-z])(?=[A-Za-z\d]*\d)[A-Za-z\d]+\b")
# Acronym: two or more consecutive uppercase letters
_ACRONYM_RE = re.compile(r"\b[A-Z]{2,}\b")
# Cardinal: plain integer/decimal (lowest priority — only claims leftover digit runs)
_CARDINAL_RE = re.compile(r"\b\d[\d,]*(?:\.\d+)?\b")

# Priority order — earlier entries claim overlapping spans first.
_NSW_MATCHERS = [
    ("Bracketed",    _BRACKET_RE),
    ("Email",        _EMAIL_RE),
    ("Hashtag",      _HASHTAG_RE),
    ("Money",        _MONEY_RE),
    ("Percentage",   _PERCENT_RE),
    ("Telephone",    _PHONE_RE),
    ("Time",         _TIME_RE),
    ("Date",         _DATE_RE),
    ("Number range", _NRANGE_RE),
    ("Ordinal",      _ORDINAL_RE),
    ("LSEQ",         _LSEQ_RE),
    ("SPLT",         _SPLT_RE),
    ("Acronym",      _ACRONYM_RE),
    ("Cardinal",     _CARDINAL_RE),
]

# Abbreviation detection uses spaCy tokens (see Pass 2), not this scanner.
# spaCy keeps known abbreviations as single tokens with the trailing period attached
# (e.g. "Dr.", "Art.", "vs."), while sentence-final periods are split off as
# standalone "." punctuation tokens.  We further exclude pure-uppercase dotted
# tokens (e.g. "A.K.", "U.S.A.") since those belong to LSEQ.
_PURE_CAPS_DOTTED = re.compile(r"^[A-Z.]+$")


def _scan_nsw(text: str, freq: dict[str, dict[str, int]]) -> None:
    """
    Scan `text` for all NSW categories, resolving overlaps by matcher priority,
    and increment freq[category][token].  Each character position is claimed by at
    most one (highest-priority) matcher.
    """
    claimed: list[tuple[int, int]] = []   # accepted spans (start, end)

    def overlaps(s: int, e: int) -> bool:
        return any(s < ce and cs < e for cs, ce in claimed)

    for label, pattern in _NSW_MATCHERS:
        for m in pattern.finditer(text):
            s, e = m.start(), m.end()
            if overlaps(s, e):
                continue
            claimed.append((s, e))
            freq[label][m.group().strip()] += 1


# ── Reusable single-span classifier (shared with --align-and-review2) ─────────

# Full ordered NSW category list: the 13 regex matchers (priority order) + the
# spaCy-detected Abbreviation category.
NSW_CATEGORIES: list[str] = [label for label, _ in _NSW_MATCHERS] + ["Abbreviation"]


def classify_nsw_span(text: str, abbrev_set: set | None = None) -> str | None:
    """
    Classify a span of text into ONE NSW category, mirroring the priority-ordered
    scan used by --content-stats:

      1. the 13 regex matchers in _NSW_MATCHERS order — first that matches wins;
      2. else, if any de-punctuated token is a known abbreviation → "Abbreviation";
      3. else None (caller may bucket the span as "Other").

    `abbrev_set` is a set of lowercase short forms (with and without a trailing
    '.'), as produced by align-and-review's load_abbrev_set().
    """
    for label, pattern in _NSW_MATCHERS:
        if pattern.search(text):
            return label
    if abbrev_set:
        for tok in text.split():
            base = tok.strip().lower().rstrip(".")
            if base and (base in abbrev_set or f"{base}." in abbrev_set):
                return "Abbreviation"
    return None


# ── Transcript parsing ────────────────────────────────────────────────────────

def _parse_transcript(path: Path) -> tuple[list[str], list[str]]:
    """
    Parse a raw transcript file (output of --initial-transcript-extraction).
    Returns (speakers, utterances) — parallel lists, one entry per turn.
    """
    speakers:   list[str] = []
    utterances: list[str] = []
    cur_spk:   str | None = None
    cur_lines: list[str]  = []

    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        m = _SPEAKER_RE.match(line)
        if m:
            if cur_spk is not None and cur_lines:
                speakers.append(cur_spk)
                utterances.append(" ".join(cur_lines))
            cur_spk   = m.group(1).strip()
            cur_lines = []
        elif line and cur_spk is not None:
            cur_lines.append(line)

    if cur_spk is not None and cur_lines:
        speakers.append(cur_spk)
        utterances.append(" ".join(cur_lines))

    return speakers, utterances


# ── spaCy sentence helpers ────────────────────────────────────────────────────

def _word_count(sent) -> int:
    return sum(1 for t in sent if not t.is_punct and not t.is_space)


def _is_partial(sent) -> bool:
    """Sentence is partial if it trails off with '...' or has no verb/aux root."""
    if _ELLIPSIS_RE.search(sent.text.strip()):
        return True
    return not any(t.dep_ == "ROOT" and t.pos_ in ("VERB", "AUX") for t in sent)


# ── Excel style constants ─────────────────────────────────────────────────────

_HDR_FONT   = Font(bold=True, color="FFFFFF")
_HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL   = PatternFill("solid", fgColor="D6E4F0")
_TITLE_FONT = Font(bold=True, size=11)
_CENTER     = Alignment(horizontal="center", vertical="center")
_LEFT       = Alignment(horizontal="left",   vertical="center")
_THIN       = Side(style="thin", color="AAAAAA")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(ws, row: int, col: int, value, align=None) -> None:
    c = ws.cell(row, col, value)
    c.font = _HDR_FONT; c.fill = _HDR_FILL
    c.border = _BORDER; c.alignment = align or _CENTER


def _data(ws, row: int, col: int, value, fill=None, align=None) -> None:
    c = ws.cell(row, col, value)
    c.border = _BORDER; c.alignment = align or _LEFT
    if fill:
        c.fill = fill


def _alt(row: int):
    return _ALT_FILL if row % 2 == 0 else None


def _set_col_widths(ws, widths: list[int]) -> None:
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Sheet 1 — Overview ────────────────────────────────────────────────────────

def _build_sheet1(
    ws,
    n_files: int,
    n_speakers: int,
    n_complete: int,
    n_partial: int,
    avg_c: float,
    avg_p: float,
    max_c: int,
    max_p: int,
) -> None:
    ws.title = "Overview"
    total = n_complete + n_partial

    # ── Table A: Speaker & File summary ──────────────────────────────────────
    ws.cell(1, 1, "SPEAKER & FILE SUMMARY").font = _TITLE_FONT
    for col, h in enumerate(["Metric", "Value"], 1):
        _hdr(ws, 2, col, h)
    for i, (label, val) in enumerate([
        ("Transcript files processed", n_files),
        ("Unique raw speaker names",    n_speakers),
    ], 3):
        f = _alt(i)
        _data(ws, i, 1, label, f, _LEFT)
        _data(ws, i, 2, val,   f, _CENTER)

    # ── Table B: Sentence statistics ─────────────────────────────────────────
    start = 6   # leave one blank row after Table A
    ws.cell(start, 1, "SENTENCE STATISTICS").font = _TITLE_FONT
    for col, h in enumerate(
        ["Sentence Type", "Count", "% of Total", "Avg Words", "Max Words"], 1
    ):
        _hdr(ws, start + 1, col, h)

    pct_c = round(100 * n_complete / max(total, 1), 1)
    pct_p = round(100 * n_partial  / max(total, 1), 1)

    rows = [
        ("Total sentences",       total,      100.0,  None,             None),
        ("Logically complete",     n_complete, pct_c,  round(avg_c, 1),  max_c),
        ("Partial / fragmentary",  n_partial,  pct_p,  round(avg_p, 1),  max_p),
    ]
    for i, (label, cnt, pct, avg, mx) in enumerate(rows, start + 2):
        f = _alt(i)
        _data(ws, i, 1, label, f, _LEFT)
        _data(ws, i, 2, cnt,   f, _CENTER)
        _data(ws, i, 3, pct,   f, _CENTER)
        _data(ws, i, 4, avg,   f, _CENTER)
        _data(ws, i, 5, mx,    f, _CENTER)

    _set_col_widths(ws, [36, 12, 14, 12, 12])


# ── Sheet 2 — Category Counts ─────────────────────────────────────────────────

def _build_sheet2(ws, n_total: int, cats: dict) -> None:
    ws.title = "Category Counts"
    headers = ["Category", "Sentence Count", "% of Total Sentences"]
    for col, (h, w) in enumerate(zip(headers, [46, 18, 22]), 1):
        _hdr(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    rows = [
        ("Containing numerals (any digit)",        cats["numeral"]),
        ("Containing dates",                        cats["date"]),
        ("Containing clock times",                  cats["time"]),
        ("Containing years (1700–2099)",            cats["year"]),
        ("Containing measuring units (km, kg, …)", cats["measure"]),
    ]
    for i, (label, cnt) in enumerate(rows, 2):
        f = _alt(i)
        _data(ws, i, 1, label,                                      f, _LEFT)
        _data(ws, i, 2, cnt,                                        f, _CENTER)
        _data(ws, i, 3, round(100 * cnt / max(n_total, 1), 1),     f, _CENTER)

    ws.freeze_panes = "A2"


# ── Sheet 3 — Character Frequency ─────────────────────────────────────────────

def _build_sheet3(ws, char_freq: dict) -> None:
    ws.title = "Character Frequency"
    headers = ["Rank", "Char", "Unicode", "Unicode Name", "Count", "% of Total"]
    for col, (h, w) in enumerate(zip(headers, [7, 8, 10, 36, 12, 12]), 1):
        _hdr(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    total = sum(char_freq.values())
    for i, (ch, cnt) in enumerate(
        sorted(char_freq.items(), key=lambda x: -x[1]), 2
    ):
        f = _alt(i)
        try:
            uname = unicodedata.name(ch)
        except ValueError:
            uname = "CONTROL"
        display = repr(ch) if ch in (" ", "\t") else ch
        _data(ws, i, 1, i - 1,                              f, _CENTER)
        _data(ws, i, 2, display,                             f, _CENTER)
        _data(ws, i, 3, f"U+{ord(ch):04X}",                 f, _CENTER)
        _data(ws, i, 4, uname,                               f, _LEFT)
        _data(ws, i, 5, cnt,                                 f, _CENTER)
        _data(ws, i, 6, round(100 * cnt / max(total, 1), 3), f, _CENTER)

    ws.freeze_panes = "A2"


# ── Sheet 4 — Word Frequency ──────────────────────────────────────────────────

def _build_sheet4(ws, word_freq: dict) -> None:
    ws.title = "Word Frequency"
    headers = ["Rank", "Word", "Frequency", "% of Total"]
    for col, (h, w) in enumerate(zip(headers, [8, 40, 14, 14]), 1):
        _hdr(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    total   = sum(word_freq.values())
    ranked  = sorted(word_freq.items(), key=lambda x: -x[1])
    for i, (word, cnt) in enumerate(ranked, 2):
        f = _alt(i)
        _data(ws, i, 1, i - 1,                               f, _CENTER)
        _data(ws, i, 2, word,                                 f, _LEFT)
        _data(ws, i, 3, cnt,                                  f, _CENTER)
        _data(ws, i, 4, round(100 * cnt / max(total, 1), 4), f, _CENTER)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(ranked) + 1}"


# ── Sheet 5 — NSW Frequency ───────────────────────────────────────────────────

def _build_sheet5(ws, nsw_rows: list) -> None:
    ws.title = "NSW Frequency"
    headers = ["Token", "Category", "Frequency"]
    for col, (h, w) in enumerate(zip(headers, [28, 16, 12]), 1):
        _hdr(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, (tok, cat, cnt) in enumerate(nsw_rows, 2):
        f = _alt(i)
        _data(ws, i, 1, tok, f, _LEFT)
        _data(ws, i, 2, cat, f, _CENTER)
        _data(ws, i, 3, cnt, f, _CENTER)

    ws.freeze_panes = "A2"
    if nsw_rows:
        ws.auto_filter.ref = f"A1:C{len(nsw_rows) + 1}"


# ── Main entry point ──────────────────────────────────────────────────────────

def run_content_stats(transcript_dir: str, output_dir: str, tag: str) -> None:
    try:
        import spacy
    except ImportError:
        raise SystemExit("ERROR: spacy not installed.  Run: pip install spacy")

    print("  Loading spaCy model en_core_web_sm...")
    try:
        nlp = spacy.load("en_core_web_sm")
    except OSError:
        raise SystemExit(
            "ERROR: spaCy model not found.\n"
            "Install with:  python -m spacy download en_core_web_sm"
        )

    txts = sorted(Path(transcript_dir).glob("transcript_*.txt"))
    if not txts:
        raise SystemExit(
            f"ERROR: No transcript_*.txt files found in {transcript_dir}"
        )
    print(f"  Found {len(txts)} transcript files. Parsing...")

    # ── Pass 1: collect utterances and surface-level counts ───────────────────
    unique_speakers: set[str]  = set()
    all_utterances:  list[str] = []

    char_freq: dict[str, int] = defaultdict(int)
    word_freq: dict[str, int] = defaultdict(int)

    # NSW frequency: category → token → count
    nsw_freq: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for path in txts:
        speakers, utterances = _parse_transcript(path)
        unique_speakers.update(speakers)

        for utt in utterances:
            all_utterances.append(utt)

            # Character frequency (spoken text, skip newlines)
            for ch in utt:
                if ch not in ("\n", "\r"):
                    char_freq[ch] += 1

            # Word frequency (lowercased, punctuation stripped)
            for tok in utt.split():
                word = _STRIP_PUNCT.sub("", tok.lower())
                if word:
                    word_freq[word] += 1

            # NSW tokens — priority-ordered scan resolves overlaps
            _scan_nsw(utt, nsw_freq)

    print(
        f"  Utterances: {len(all_utterances):,}  |  "
        f"Unique speakers: {len(unique_speakers)}"
    )
    print("  Running spaCy pipeline (may take a minute for large corpora)...")

    # ── Pass 2: spaCy sentence analysis + abbreviation detection ─────────────
    n_complete = n_partial = 0
    complete_wc: list[int] = []
    partial_wc:  list[int] = []
    cats: dict[str, int] = defaultdict(int)

    for doc in nlp.pipe(all_utterances, batch_size=64, disable=["ner"]):
        # Abbreviation: spaCy keeps known abbreviations as one token with trailing period.
        # Exclude pure-uppercase dotted tokens (those belong to LSEQ).
        for token in doc:
            if (token.text.endswith(".")
                    and len(token.text) > 1
                    and not _PURE_CAPS_DOTTED.match(token.text)):
                nsw_freq["Abbreviation"][token.text] += 1
        for sent in doc.sents:
            wc = _word_count(sent)
            if wc < 2:
                continue
            txt = sent.text
            if _is_partial(sent):
                n_partial += 1
                partial_wc.append(wc)
            else:
                n_complete += 1
                complete_wc.append(wc)
            if _NUMERAL_RE.search(txt): cats["numeral"] += 1
            if _DATE_RE.search(txt):    cats["date"]    += 1
            if _TIME_RE.search(txt):    cats["time"]    += 1
            if _YEAR_RE.search(txt):    cats["year"]    += 1
            if _MEASURE_RE.search(txt): cats["measure"] += 1

    n_total = n_complete + n_partial
    avg_c   = sum(complete_wc) / max(len(complete_wc), 1)
    avg_p   = sum(partial_wc)  / max(len(partial_wc),  1)
    max_c   = max(complete_wc) if complete_wc else 0
    max_p   = max(partial_wc)  if partial_wc  else 0

    print(
        f"  Sentences: {n_total:,}  "
        f"(complete: {n_complete:,}, partial: {n_partial:,})"
    )

    # ── Assemble NSW rows (all categories, sorted by frequency desc) ──────────
    nsw_rows: list[tuple] = [
        (tok, category, cnt)
        for category, tokens in nsw_freq.items()
        for tok, cnt in tokens.items()
    ]
    nsw_rows.sort(key=lambda r: -r[2])

    # ── Build workbook ────────────────────────────────────────────────────────
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f"{tag}_transcript_content_analysis.xlsx")

    wb = openpyxl.Workbook()
    _build_sheet1(
        wb.active, len(txts), len(unique_speakers),
        n_complete, n_partial, avg_c, avg_p, max_c, max_p,
    )
    _build_sheet2(wb.create_sheet(), n_total, cats)
    _build_sheet3(wb.create_sheet(), dict(char_freq))
    _build_sheet4(wb.create_sheet(), dict(word_freq))
    _build_sheet5(wb.create_sheet(), nsw_rows)

    wb.save(out_path)
    print(f"  Written: {out_path}")
