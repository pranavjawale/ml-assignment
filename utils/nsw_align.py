"""
utils/nsw_align.py

V1 ↔ V3 token alignment + Flint-taxonomy NSW classification + Excel writers.
Imported by analysis-pipeline.py for the --align-and-review[2] modes.

Aligns two token sequences per speaker turn using word-level edit distance
(difflib.SequenceMatcher) and classifies each difference into the Flint et al.
(2017) NSW taxonomy. The Excel writers build one sheet per NSW category.

Data-sheet columns:
  Written (V1)    -- written token(s) that differ from V3
  Spoken (V3)     -- corresponding V3 token(s) after expansion
  V1 Context      -- up to CONTEXT_WIN tokens around the NSW (NSW in red)
  V3 Context      -- same span in V3 (expansion in blue)
  Turn            -- speaker ID + turn index for traceability
"""

import re
import difflib
from pathlib import Path

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont

CONTEXT_WIN    = 5   # tokens of context on each side of NSW
MAX_V1_NSW_LEN = 12  # V1 spans longer than this are SequenceMatcher alignment artifacts

# ── NSW categories in display order ───────────────────────────────────────────
# Each entry: (sheet_name, description shown in SUMMARY sheet)
NSW_CATS = [
    ("ALPHA-EXPN",
     "Abbreviation / acronym expanded to full words  (Dr. -> doctor, CJI -> chief justice of india)"),
    ("ALPHA-LSEQ",
     "Letter sequence unchanged between V1 and V3  (spaced initials like 'B V' -- rarely differs)"),
    ("ALPHA-WDLK",
     "Word-like token transformed  (Hon ble -> honourable, para -> paragraph)"),
    ("NUM",
     "Cardinal number -> spoken words  (370 -> three seventy, 21 -> twenty one)"),
    ("NORD",
     "Ordinal number -> spoken ordinal  (1st -> first, 15th -> fifteenth)"),
    ("NYER",
     "Year -> spoken year  (1950 -> nineteen fifty, 2023 -> twenty twenty three)"),
    ("NRANGE",
     "Number range with hyphen  (25-30) -- hyphen NOT yet replaced with 'to'"),
    ("MONEY",
     "Monetary expression  (Rs., INR) -- Rs. currently not in TSV, not expanded"),
    ("PRCT",
     "Percentage  (23%, 23.5%) -- '%' not yet expanded to 'percent'"),
    ("NDIG",
     "Digit string / decimal (case numbers, page refs)  -- may be wrongly split/expanded"),
    ("NTIME",
     "Time expression  (9:30, 17:10) -- digit parts expanded but colon kept"),
    ("NDATE",
     "Date expression  (15-08-1947, 26/11) -- digits expanded but structure not parsed"),
    ("SPLT",
     "Fused alphanumeric  (39B, 124A, 31C) -- NOT yet split or expanded in V3"),
    ("PAREN",
     "Parenthetical reference stripped in V3  (7 ( 2 ) -> seven two)"),
    ("MISC-NONE",
     "Unclassified difference -- review to find new patterns or false positives"),
]
CAT_NAMES = [c for c, _ in NSW_CATS]
CAT_DESC  = {c: d for c, d in NSW_CATS}

# ── Regex helpers ──────────────────────────────────────────────────────────────
_PURE_NUM  = re.compile(r'^\d+$')
_ORDINAL   = re.compile(r'^\d{1,4}(st|nd|rd|th)$', re.I)
_DECIMAL   = re.compile(r'^\d+\.\d+$')
_SPLT      = re.compile(r'^\d+[A-Za-z]+$|^[A-Za-z]+\d+$')
_PERCENT   = re.compile(r'^\d*%$|^\d+\.\d+%$')
_DATEPAT   = re.compile(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$')
_TIMEPAT   = re.compile(r'^\d{1,2}:\d{2}(am|pm)?$', re.I)
_RANGE_SEP = re.compile(r'^-$')
_SPK_LINE  = re.compile(r'^(spk_\d{4}):$')
# Trailing ?! on a V1 token: V1 doesn't space these but V3's f2 does.
# Used only for alignment pre-processing (not for display).
_TRAIL_QX  = re.compile(r'^(.+?)([?!]+)$')


# ── Load abbreviation set from TSV ────────────────────────────────────────────

def load_abbrev_set(tsv_path: Path) -> set:
    """Return lowercase set of all short forms (with and without trailing period)."""
    result = set()
    for line in Path(tsv_path).read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        parts = line.split('\t')
        if not parts:
            continue
        short = parts[0].strip()
        lo = short.lower()
        result.add(lo)
        if lo.endswith('.'):
            result.add(lo[:-1])
    return result


# ── Transcript parser ─────────────────────────────────────────────────────────

def parse_transcript(path: Path) -> list:
    """Return list of (spk_id, body_str) from a V1 or V3 transcript file."""
    turns, cur_spk, cur_body = [], None, []
    for line in Path(path).read_text(encoding='utf-8').splitlines():
        line = line.rstrip()
        m = _SPK_LINE.match(line)
        if m:
            if cur_spk is not None and cur_body:
                turns.append((cur_spk, ' '.join(cur_body)))
            cur_spk, cur_body = m.group(1), []
        elif not line:
            if cur_spk is not None and cur_body:
                turns.append((cur_spk, ' '.join(cur_body)))
                cur_spk, cur_body = None, []
        elif cur_spk is not None:
            cur_body.append(line.strip())
    if cur_spk and cur_body:
        turns.append((cur_spk, ' '.join(cur_body)))
    return turns


# ── NSW classifier ────────────────────────────────────────────────────────────

def classify(v1_orig: list, v3_toks: list, abbrev_set: set) -> str:
    """Classify a (V1 span, V3 span) difference into an NSW category string."""

    # Pure deletion from V1 (V3 is empty)
    if not v1_orig:
        return "MISC-NONE"
    if not v3_toks:
        if all(t in ('(', ')') for t in v1_orig):
            return "PAREN"
        return "MISC-NONE"

    # Any parenthesis in the V1 span
    if '(' in v1_orig or ')' in v1_orig:
        return "PAREN"

    # ── Single-token V1 ────────────────────────────────────────────────────────
    if len(v1_orig) == 1:
        tok = v1_orig[0]
        if _ORDINAL.match(tok):      return "NORD"
        if _SPLT.match(tok):         return "SPLT"
        if _PERCENT.match(tok):      return "PRCT"
        if _TIMEPAT.match(tok):      return "NTIME"
        if _DATEPAT.match(tok):      return "NDATE"
        if _DECIMAL.match(tok):      return "NDIG"
        if _PURE_NUM.match(tok):
            n = int(tok)
            return "NYER" if 1800 <= n <= 2099 else "NUM"
        tok_lo = tok.lower()
        if tok_lo in abbrev_set:
            # Distinguish EXPN (title/acronym expanded) vs WDLK (wordlike abbreviated form)
            # Word-like entries tend to be lowercase already (para, govt, etc.)
            if tok.isupper() or tok.endswith('.') or tok[0].isupper():
                return "ALPHA-EXPN"
            return "ALPHA-WDLK"
        if re.match(r'^[A-Z]{2,}$', tok):
            return "ALPHA-LSEQ"
        return "MISC-NONE"

    # ── Multi-token V1 ─────────────────────────────────────────────────────────

    # Dotted abbreviation spaced by V1: ["Dr", "."] or ["No", "."] etc.
    if len(v1_orig) == 2 and v1_orig[1] == '.':
        if v1_orig[0].lower() in abbrev_set:
            return "ALPHA-EXPN"

    # Territorial / ampersand abbreviation: J&K -> Jammu and Kashmir
    if any('&' in t for t in v1_orig):
        return "ALPHA-EXPN"

    # Multi-word short forms: "Hon ble" -> honourable
    joined_lo  = ' '.join(v1_orig).lower()
    joined_nsp = joined_lo.replace(' ', '')
    if joined_lo in abbrev_set or joined_nsp in abbrev_set:
        return "ALPHA-EXPN"

    # Number range: NUM - NUM
    if (len(v1_orig) == 3
            and _RANGE_SEP.match(v1_orig[1])
            and _PURE_NUM.match(v1_orig[0])
            and _PURE_NUM.match(v1_orig[2])):
        return "NRANGE"

    # Date: digit / sep / digit / sep / digit (e.g. 01 - 01 - 64 when tokenised)
    if all(_PURE_NUM.match(t) or t in ('-', '/') for t in v1_orig):
        seps = [t for t in v1_orig if t in ('-', '/')]
        if len(seps) >= 2:
            return "NDATE"
        if seps:
            return "NRANGE"
        return "NUM"

    # Currency prefix: Rs . 500 or Rs. 10 lakhs
    if v1_orig[0].lower() in ('rs', 'rs.'):
        return "MONEY"

    # Fallback scan: classify by the first recognizable token in the span
    for tok in v1_orig:
        tok_lo = tok.lower()
        if tok_lo in abbrev_set:
            return "ALPHA-EXPN"
        if _ORDINAL.match(tok):
            return "NORD"
        if _PURE_NUM.match(tok):
            n = int(tok)
            return "NYER" if 1800 <= n <= 2099 else "NUM"
        if _DECIMAL.match(tok):
            return "NDIG"
        if _PERCENT.match(tok):
            return "PRCT"

    return "MISC-NONE"


# ── Token alignment ───────────────────────────────────────────────────────────

def _expand_v1_for_alignment(v1_toks: list) -> tuple:
    """
    Pre-split trailing ?! off V1 tokens and lowercase, to match V3's f2 spacing.
    V1 already splits . from words (via _space_inline_punct), but leaves ?! fused.
    Returns (expanded_lower_toks, idx_map) where idx_map[i] = original V1 index.
    """
    expanded, idx_map = [], []
    for i, tok in enumerate(v1_toks):
        tl = tok.lower()
        m  = _TRAIL_QX.match(tl)
        if m:
            expanded.append(m.group(1))   # word part
            expanded.append(m.group(2))   # ?! part
            idx_map.extend([i, i])
        else:
            expanded.append(tl)
            idx_map.append(i)
    return expanded, idx_map


def align_turn(v1_toks: list, v3_toks: list) -> list:
    """
    Return list of difference dicts {v1_orig, v3, i1, i2, j1, j2} for non-equal spans.
    Pre-splits trailing ?! off V1 tokens for alignment (V3 does this via f2), then maps
    expanded indices back to original V1 token indices for display.
    Skips diffs where the original V1 span exceeds MAX_V1_NSW_LEN — those are alignment
    artifacts produced by SequenceMatcher on long turns, not real NSWs.
    """
    v1_exp, idx_map = _expand_v1_for_alignment(v1_toks)
    sm = difflib.SequenceMatcher(None, v1_exp, v3_toks, autojunk=False)
    diffs = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == 'equal':
            continue
        # Map expanded indices back to original V1 token positions
        orig_i1 = idx_map[i1] if i1 < len(idx_map) else len(v1_toks)
        orig_i2 = (idx_map[i2 - 1] + 1) if i2 > i1 else orig_i1
        v1_span = v1_toks[orig_i1:orig_i2]
        if len(v1_span) > MAX_V1_NSW_LEN:
            continue  # alignment artifact — skip
        diffs.append({
            'v1_orig': v1_span,
            'v3':      v3_toks[j1:j2],
            'i1': orig_i1, 'i2': orig_i2,
            'j1': j1, 'j2': j2,
        })
    return diffs


def get_context(v1_toks: list, i1: int, i2: int) -> tuple:
    before = v1_toks[max(0, i1 - CONTEXT_WIN): i1]
    after  = v1_toks[i2: min(len(v1_toks), i2 + CONTEXT_WIN)]
    return ' '.join(before), ' '.join(after)


def get_v3_context(v3_toks: list, j1: int, j2: int) -> tuple:
    before = v3_toks[max(0, j1 - CONTEXT_WIN): j1]
    after  = v3_toks[j2: min(len(v3_toks), j2 + CONTEXT_WIN)]
    return ' '.join(before), ' '.join(after)


# ── Excel writers ─────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="2E4057")
_HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_DAT_FONT = Font(name="Calibri", size=10)
_ALT_FILL = PatternFill("solid", fgColor="EEF2F7")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_HEADERS  = ["Written (V1)", "Spoken (V3)", "V1 Context  (NSW in red)", "V3 Context  (expansion in blue)", "Turn"]
_COL_W    = [24, 38, 58, 58, 14]

_CTX_BLACK = InlineFont(rFont="Calibri", sz=10, color="000000")
_CTX_RED   = InlineFont(rFont="Calibri", sz=10, color="FF0000", b=True)
_CTX_BLUE  = InlineFont(rFont="Calibri", sz=10, color="0070C0", b=True)


def _rich_context(before: str, nsw: str, after: str, hi_font=None) -> CellRichText:
    """Build a cell value with the NSW/expansion highlighted and context in black."""
    if hi_font is None:
        hi_font = _CTX_RED
    parts = []
    if before:
        parts.append(TextBlock(_CTX_BLACK, before + " "))
    parts.append(TextBlock(hi_font, nsw or "(empty)"))
    if after:
        parts.append(TextBlock(_CTX_BLACK, " " + after))
    return CellRichText(*parts)


def _write_data_sheet(wb, cat: str, rows: list) -> None:
    ws = wb.create_sheet(title=cat)
    ws.row_dimensions[1].height = 18
    for ci, (hdr, w) in enumerate(zip(_HEADERS, _COL_W), start=1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}1"

    for ri, row in enumerate(rows, start=2):
        v1_str, v3_str, bef_v1, aft_v1, bef_v3, aft_v3, label = row
        use_alt = ri % 2 == 0

        # Col 1: Written (V1)
        c1 = ws.cell(row=ri, column=1, value=v1_str)
        c1.font = _DAT_FONT; c1.alignment = _WRAP_TOP
        if use_alt: c1.fill = _ALT_FILL

        # Col 2: Spoken (V3)
        c2 = ws.cell(row=ri, column=2, value=v3_str)
        c2.font = _DAT_FONT; c2.alignment = _WRAP_TOP
        if use_alt: c2.fill = _ALT_FILL

        # Col 3: V1 context with NSW in red
        c3 = ws.cell(row=ri, column=3, value=_rich_context(bef_v1, v1_str, aft_v1, _CTX_RED))
        c3.alignment = _WRAP_TOP
        if use_alt: c3.fill = _ALT_FILL

        # Col 4: V3 context with expansion in blue
        c4 = ws.cell(row=ri, column=4, value=_rich_context(bef_v3, v3_str, aft_v3, _CTX_BLUE))
        c4.alignment = _WRAP_TOP
        if use_alt: c4.fill = _ALT_FILL

        # Col 5: Turn
        c5 = ws.cell(row=ri, column=5, value=label)
        c5.font = _DAT_FONT; c5.alignment = _WRAP_TOP
        if use_alt: c5.fill = _ALT_FILL

    if not rows:
        c = ws.cell(row=2, column=1,
                    value="(no instances found -- this category is not yet handled or not present)")
        c.font = Font(italic=True, color="888888", name="Calibri", size=10)


def _write_summary(wb, counts: dict, fname: str) -> None:
    ws = wb.create_sheet(title="SUMMARY", index=0)
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 72
    ws.cell(row=1, column=1, value="File:").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=1, column=2, value=fname).font   = Font(name="Calibri", size=10)
    for ci, hdr in enumerate(["Category", "Count", "Description / Notes"], start=1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = Alignment(
            horizontal="center" if ci <= 2 else "left", vertical="center")
    for ri, (cat, desc) in enumerate(NSW_CATS, start=4):
        ws.cell(row=ri, column=1, value=cat).font = Font(bold=True, name="Calibri", size=10)
        n = ws.cell(row=ri, column=2, value=counts.get(cat, 0))
        n.font      = _DAT_FONT
        n.alignment = Alignment(horizontal="center")
        d = ws.cell(row=ri, column=3, value=desc)
        d.font = Font(name="Calibri", size=10, color="333333")
