"""
Analysis pipeline — main entry point.

Outputs under analysis-outputs/ (all prefixed with <tag>):
  <tag>_pdf-analysis-summary.txt
  <tag>_transcript_content_analysis.xlsx
  <tag>_speaker-name-analysis.xlsx   (SHEET1 stats, SHEET2 matrix, SHEET3 dedup review)
  <tag>_spk-id-mapping.txt
  <tag>_norm_trans_word_freq.xlsx
  <tag>_transcript_original/transcript_NNN.txt              — speaker names as-is from PDF
  <tag>_spkid_mapped_transcript_v1/transcript_NNN.txt       — spk_NNNN labels, V1 cleanup
  <tag>_spkid_mapped_norm_transcript_v3/transcript_NNN.txt  — spk_NNNN labels, V3 spoken form

Usage:
    # PDF structural metadata
    python analysis-pipeline.py --pdf-summary --pdf-dir <pdf-dir> --tag <label>

    # Raw transcripts → analysis-outputs/<label>_transcript_original/
    python analysis-pipeline.py --extract-trans --pdf-dir <pdf-dir> --tag <label>

    # spaCy content statistics (reads <label>_transcript_original/)
    python analysis-pipeline.py --content-stats --tag <label>

    # Speaker name deduplication only (reads <label>_transcript_original/)
    python analysis-pipeline.py --dedup-names --tag <label>

    # Full pipeline from PDFs → normalized transcripts + analysis outputs
    python analysis-pipeline.py --get-norm-trans --pdf-dir <pdf-dir> --tag <label>

    # Analysis-only on already-extracted transcripts (dedup + content-stats)
    python analysis-pipeline.py --get-norm-trans --tag <label>

    # NSW review (main): transcript_original vs V3, content-stats categories
    #   → <tag>_nsw-review-<id>.xlsx
    python analysis-pipeline.py --align-and-review --tag <label> --file-id 001,003

    # NSW review (alt): V1 vs V3, same content-stats categories → <tag>_nsw-review2-<id>.xlsx
    python analysis-pipeline.py --align-and-review2 --tag <label> --file-id 001,003
"""

import argparse
import os
import sys
import re
import time
from collections import defaultdict
from pathlib import Path

import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from utils.extractor           import PdfExtractor
from utils.transcript_extractor import TranscriptExtractor
from utils.deduplicator        import SpeakerDeduplicator
from utils.reporters      import write_pdf_analysis_summary
from utils.excel_builder  import build_speaker_excel, write_spk_id_mapping
from utils.aliases        import RAW_ALIASES
from utils.content_stats  import run_content_stats, classify_nsw_span, NSW_CATEGORIES
from utils.nsw_align      import (
    align_turn, get_context, get_v3_context, load_abbrev_set, _write_data_sheet,
)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Output directory: analysis-outputs/ alongside this script
OUT_DIR = os.path.join(SCRIPT_DIR, "analysis-outputs")

_RE_BOILERPLATE = re.compile(r"END\s+OF\s+DAY'?S?\s+PROCEEDINGS", re.IGNORECASE)

_CLEANUPS = [
    ('"',  ' '), ('"', ' '), ('"', ' '),
    ('‘', ' '), ('’', ' '),
    ('—', ' '), ('–', ' '),
    ('…', ' '), ('²', ' '),
]

_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
_CENTER   = Alignment(horizontal="center", vertical="center")
_LEFT     = Alignment(horizontal="left",   vertical="center")
_THIN     = Side(style="thin", color="AAAAAA")
_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ── Text cleanup ──────────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    for old, new in _CLEANUPS:
        text = text.replace(old, new)
    return re.sub(r' +', ' ', text).strip()


# ── Speaker resolution ────────────────────────────────────────────────────────

def resolve_speaker(raw_name: str, spk_map: dict[str, str]) -> tuple[str, bool]:
    canonical = RAW_ALIASES.get(raw_name, raw_name)
    spk_id    = spk_map.get(canonical) or spk_map.get(raw_name)
    if spk_id:
        return spk_id, True
    return canonical, False


# ── Turn extraction ───────────────────────────────────────────────────────────

def extract_turns(pdf_path: str, spk_map: dict[str, str],
                  unmapped: set[str]) -> list[tuple[str, str, str]]:
    """Return (spk_id, pdf_name, text) triples for each speaker turn."""
    doc   = fitz.open(pdf_path)
    turns: list[tuple[str, str, str]] = []
    cur_speaker:  str | None = None
    cur_pdf_name: str | None = None
    cur_parts:    list[str]  = []

    for pi in range(1, len(doc)):
        for raw_line in doc[pi].get_text("text").splitlines():
            line = raw_line.strip()
            if not line:
                continue
            if (PdfExtractor.RE_TERES.match(line)     or
                    PdfExtractor.RE_INTEGER.match(line)   or
                    PdfExtractor.RE_TIMESTAMP.match(line) or
                    PdfExtractor.RE_EMAIL.match(line)     or
                    _RE_BOILERPLATE.search(line)):
                continue
            m = PdfExtractor.RE_SPEAKER.match(line)
            if m:
                if cur_speaker is not None and cur_parts:
                    turns.append((cur_speaker, cur_pdf_name, " ".join(cur_parts)))
                raw_name            = m.group(1).strip()
                cur_speaker, found  = resolve_speaker(raw_name, spk_map)
                cur_pdf_name        = raw_name
                if not found:
                    unmapped.add(raw_name)
                cur_parts = []
                rest = m.group(2).strip()
                if rest:
                    cur_parts.append(rest)
            elif cur_speaker is not None:
                cur_parts.append(line)

    if cur_speaker is not None and cur_parts:
        turns.append((cur_speaker, cur_pdf_name, " ".join(cur_parts)))

    doc.close()
    return turns


# ── Transcript writer ─────────────────────────────────────────────────────────

def write_transcript(turns: list[tuple[str, str]], output_path: str) -> list[str]:
    """Write transcript file; returns list of word tokens in the spoken text."""
    tokens: list[str] = []
    with open(output_path, "w", encoding="utf-8") as f:
        for speaker, text in turns:
            cleaned = clean_text(text)
            if not cleaned:
                continue
            f.write(f"{speaker}:\n{cleaned}\n\n")
            tokens.extend(cleaned.split())
    return tokens


# ── Word-frequency Excel ──────────────────────────────────────────────────────

def _wf_hdr(cell, align=None):
    cell.font = _HDR_FONT; cell.fill = _HDR_FILL
    cell.border = _BORDER; cell.alignment = align or _CENTER


def _wf_cell(cell, fill=None, align=None):
    cell.border = _BORDER; cell.alignment = align or _LEFT
    if fill:
        cell.fill = fill


def write_word_freq_excel(word_freq: dict[str, int], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Word Frequency"

    headers = ["Rank", "Word", "Frequency"]
    widths  = [8, 40, 14]
    ws.row_dimensions[1].height = 24
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _wf_hdr(ws.cell(1, col, h))
        ws.column_dimensions[get_column_letter(col)].width = w

    ranked = sorted(word_freq.items(), key=lambda x: -x[1])
    for row_idx, (word, freq) in enumerate(ranked, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        _wf_cell(ws.cell(row_idx, 1, row_idx - 1), fill, _CENTER)
        _wf_cell(ws.cell(row_idx, 2, word),         fill, _LEFT)
        _wf_cell(ws.cell(row_idx, 3, freq),         fill, _CENTER)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(ranked) + 1}"
    wb.save(output_path)
    print(f"Written: {output_path}  ({len(ranked):,} unique words)")


# ── Pipeline class ────────────────────────────────────────────────────────────

class AnalysisPipeline:
    """
    Orchestrates extraction -> deduplication -> report writing -> transcripts.

    Steps:
      1. PdfExtractor.extract_all()       — reads all PDFs
      2. write_pdf_analysis_summary()     — <tag>_pdf-analysis-summary.txt
      3. SpeakerDeduplicator.run()        — clusters & assigns IDs
      4. write_spk_id_mapping()           — <tag>_spk-id-mapping.txt
      5. build_speaker_excel()            — <tag>_speaker-name-analysis.xlsx
                                            (SHEET3 = dedup review)
      6. extract_turns() + write_transcript() + utils.normalizer (V1/V3)
                                          — <tag>_transcript_original/            (raw names)
                                            <tag>_spkid_mapped_transcript_v1/     (spk_NNNN, cleaned)
                                            <tag>_spkid_mapped_norm_transcript_v3/ (spk_NNNN, spoken form)
      7. write_word_freq_excel()          — <tag>_norm_trans_word_freq.xlsx  (from V3)
    """

    def __init__(self, pdf_dir: str, output_dir: str, tag: str):
        self.pdf_dir    = pdf_dir
        self.output_dir = output_dir
        self.tag        = tag
        self.extractor:    PdfExtractor        | None = None
        self.deduplicator: SpeakerDeduplicator | None = None

    def run(self) -> None:
        t0 = time.time()
        print("\n" + "=" * 60)
        print("ANALYSIS PIPELINE")
        print("=" * 60)

        # Step 1 — extract
        print("\n[1/7] Extracting speaker data from PDFs...")
        self.extractor = PdfExtractor(self.pdf_dir).extract_all()

        # Step 2 — pdf analysis summary
        print(f"\n[2/7] Writing {self.tag}_pdf-analysis-summary.txt...")
        write_pdf_analysis_summary(
            self.extractor,
            os.path.join(self.output_dir, f"{self.tag}_pdf-analysis-summary.txt"),
        )

        # Step 3 — dedup
        print("\n[3/7] Running speaker deduplication...")
        self.deduplicator = SpeakerDeduplicator(self.extractor).run()

        # Step 4 — speaker ID mapping
        print(f"\n[4/7] Writing {self.tag}_spk-id-mapping.txt...")
        write_spk_id_mapping(
            self.deduplicator,
            os.path.join(self.output_dir, f"{self.tag}_spk-id-mapping.txt"),
        )

        # Step 5 — Excel (SHEET3 carries the dedup review)
        print(f"\n[5/7] Building {self.tag}_speaker-name-analysis.xlsx...")
        build_speaker_excel(
            self.extractor,
            self.deduplicator,
            os.path.join(self.output_dir, f"{self.tag}_speaker-name-analysis.xlsx"),
        )

        # Step 6 — transcripts: raw originals + spk_NNNN-mapped V1 / V3 (spoken form)
        # V1 / V3 normalization lives in utils/normalizer.py (imported lazily —
        # only this mode needs num2words).
        from utils import normalizer as nt

        orig_dir = os.path.join(self.output_dir, f"{self.tag}_transcript_original")
        v1_dir   = os.path.join(self.output_dir, f"{self.tag}_spkid_mapped_transcript_v1")
        v3_dir   = os.path.join(self.output_dir, f"{self.tag}_spkid_mapped_norm_transcript_v3")
        print(f"\n[6/7] Writing {self.tag}_transcript_original/, "
              f"{self.tag}_spkid_mapped_transcript_v1/, and "
              f"{self.tag}_spkid_mapped_norm_transcript_v3/...")
        for d in (orig_dir, v1_dir, v3_dir):
            os.makedirs(d, exist_ok=True)

        abbrev_tsv = os.path.join(SCRIPT_DIR, "abbrev-expansions.tsv")
        if os.path.exists(abbrev_tsv):
            table = nt.load_abbrev_table(Path(abbrev_tsv))
            print(f"  Loaded {len(table)} abbreviation entries for V3 expansion")
        else:
            print(f"  WARNING: {abbrev_tsv} not found — no abbreviation expansion in V3")
            table = []

        spk_map   = self.deduplicator.raw_spk_id
        unmapped:   set[str]       = set()
        word_freq:  dict[str, int] = defaultdict(int)

        pdfs = sorted(
            f for f in os.listdir(self.pdf_dir)
            if f.startswith("transcript_") and f.endswith(".pdf")
        )
        for fname in pdfs:
            idx  = int(fname.replace("transcript_", "").replace(".pdf", ""))
            path = os.path.join(self.pdf_dir, fname)

            turns = extract_turns(path, spk_map, unmapped)

            # raw-name original (text only lightly cleaned)
            write_transcript([(pdf_name, txt) for _, pdf_name, txt in turns],
                             os.path.join(orig_dir, f"transcript_{idx:03d}.txt"))

            # spk_NNNN-mapped turns → V1 cleanup → V3 spoken-form expansion
            spk_turns: list[tuple[str, str]] = []
            for spk, _, txt in turns:
                cleaned = clean_text(txt)
                if cleaned:
                    spk_turns.append((spk, cleaned))
            turns_v1 = nt.make_v1(spk_turns)
            turns_v3 = nt.make_v3(turns_v1, table)
            nt.write_turns(turns_v1, Path(v1_dir) / f"transcript_{idx:03d}.txt")
            nt.write_turns(turns_v3, Path(v3_dir) / f"transcript_{idx:03d}.txt")

            for _, body in turns_v3:                       # word freq from spoken form (V3)
                for tok in body.split():
                    word_freq[tok.lower()] += 1
            print(f"  [{idx:02d}] {fname}  ->  {len(turns)} turns")

        if unmapped:
            print(f"\nWARNING: {len(unmapped)} speaker name(s) not found in spk_map:")
            for name in sorted(unmapped):
                print(f"    {name}")

        # Step 7 — word frequency (from V3 spoken form)
        print(f"\n[7/7] Writing {self.tag}_norm_trans_word_freq.xlsx...")
        write_word_freq_excel(
            dict(word_freq),
            os.path.join(self.output_dir, f"{self.tag}_norm_trans_word_freq.xlsx"),
        )

        elapsed = time.time() - t0
        print(f"\n{'=' * 60}")
        print(f"Pipeline complete in {elapsed:.1f}s")
        print(f"Output directory: {self.output_dir}")
        print("=" * 60)


# ── Mode: pdf summary ────────────────────────────────────────────────────────

def run_pdf_summary(pdf_dir: str, output_dir: str, tag: str) -> None:
    """Extract PDF metadata and write <tag>_pdf-analysis-summary.txt to output_dir."""
    t0 = time.time()
    os.makedirs(output_dir, exist_ok=True)

    print("\n" + "=" * 60)
    print("PDF ANALYSIS SUMMARY")
    print("=" * 60)

    print("\n[1/2] Extracting metadata from PDFs...")
    extractor = PdfExtractor(pdf_dir).extract_all()

    print("\n[2/2] Writing pdf-analysis-summary...")
    out_path = os.path.join(output_dir, f"{tag}_pdf-analysis-summary.txt")
    write_pdf_analysis_summary(extractor, out_path)

    elapsed = time.time() - t0
    print(f"\n{'=' * 60}")
    print(f"Done in {elapsed:.1f}s")
    print(f"Output: {out_path}")
    print("=" * 60)


# ── Mode: dedup names ────────────────────────────────────────────────────────

def run_dedup_names(transcript_dir: str, output_dir: str, tag: str) -> SpeakerDeduplicator:
    """
    Speaker name deduplication only, sourced from extracted transcripts.
    Writes two outputs to output_dir:
      <tag>_spk-id-mapping.txt
      <tag>_speaker-name-analysis.xlsx   (SHEET1 stats, SHEET2 matrix, SHEET3 dedup review)
    Returns the SpeakerDeduplicator (so callers can reuse its raw_spk_id mapping).
    """
    t0 = time.time()
    os.makedirs(output_dir, exist_ok=True)

    print("\n" + "=" * 60)
    print("SPEAKER NAME DEDUPLICATION")
    print("=" * 60)

    print("\n[1/4] Reading speaker data from transcripts...")
    extractor = TranscriptExtractor(transcript_dir).extract_all()

    print("\n[2/4] Running speaker deduplication...")
    deduplicator = SpeakerDeduplicator(extractor).run()

    print(f"\n[3/4] Writing {tag}_spk-id-mapping.txt...")
    write_spk_id_mapping(
        deduplicator,
        os.path.join(output_dir, f"{tag}_spk-id-mapping.txt"),
    )

    print(f"\n[4/4] Building {tag}_speaker-name-analysis.xlsx...")
    build_speaker_excel(
        extractor,
        deduplicator,
        os.path.join(output_dir, f"{tag}_speaker-name-analysis.xlsx"),
    )

    elapsed = time.time() - t0
    print(f"\n{'=' * 60}")
    print(f"Done in {elapsed:.1f}s")
    print(f"Output directory: {output_dir}")
    print("=" * 60)
    return deduplicator


# ── Normalized transcripts (V1 / V3) from raw-name transcripts ───────────────

_ORIG_SPEAKER_RE = re.compile(r"^([A-Z][A-Z0-9 .\-']+):\s*$")


def _parse_orig_turns(path: str) -> list[tuple[str, str]]:
    """Parse a raw-name transcript file into (raw_name, body) turns."""
    turns: list[tuple[str, str]] = []
    cur_spk:  str | None = None
    cur_body: list[str]  = []
    with open(path, encoding="utf-8") as fh:
        for raw in fh:
            line = raw.strip()
            m = _ORIG_SPEAKER_RE.match(line)
            if m:
                if cur_spk is not None and cur_body:
                    turns.append((cur_spk, " ".join(cur_body)))
                cur_spk, cur_body = m.group(1).strip(), []
            elif line and cur_spk is not None:
                cur_body.append(line)
    if cur_spk is not None and cur_body:
        turns.append((cur_spk, " ".join(cur_body)))
    return turns


def write_norm_transcripts(transcript_dir: str, output_dir: str, tag: str,
                           spk_map: dict[str, str]) -> None:
    """
    Relabel the raw-name transcripts in transcript_dir to spk_NNNN (via spk_map),
    then write V1 (cleanup) and V3 (spoken form) versions plus the V3 word frequency.
    Reuses the V1/V3 logic from utils/normalizer.py.
    """
    from utils import normalizer as nt

    t0 = time.time()
    print("\n" + "=" * 60)
    print("NORMALIZED TRANSCRIPTS (V1 / V3)")
    print("=" * 60)

    v1_dir = os.path.join(output_dir, f"{tag}_spkid_mapped_transcript_v1")
    v3_dir = os.path.join(output_dir, f"{tag}_spkid_mapped_norm_transcript_v3")
    os.makedirs(v1_dir, exist_ok=True)
    os.makedirs(v3_dir, exist_ok=True)

    abbrev_tsv = os.path.join(SCRIPT_DIR, "abbrev-expansions.tsv")
    if os.path.exists(abbrev_tsv):
        table = nt.load_abbrev_table(Path(abbrev_tsv))
        print(f"  Loaded {len(table)} abbreviation entries for V3 expansion")
    else:
        print(f"  WARNING: {abbrev_tsv} not found — no abbreviation expansion in V3")
        table = []

    txts = sorted(Path(transcript_dir).glob("transcript_*.txt"))
    unmapped:  set[str]       = set()
    word_freq: dict[str, int] = defaultdict(int)

    for path in txts:
        spk_turns: list[tuple[str, str]] = []
        for raw_name, body in _parse_orig_turns(str(path)):
            spk, found = resolve_speaker(raw_name, spk_map)
            if not found:
                unmapped.add(raw_name)
            spk_turns.append((spk, body))

        turns_v1 = nt.make_v1(spk_turns)
        turns_v3 = nt.make_v3(turns_v1, table)
        nt.write_turns(turns_v1, Path(v1_dir) / path.name)
        nt.write_turns(turns_v3, Path(v3_dir) / path.name)

        for _, b in turns_v3:                          # word freq from spoken form (V3)
            for tok in b.split():
                word_freq[tok.lower()] += 1
        print(f"  {path.name}  ->  {len(spk_turns)} turns")

    if unmapped:
        print(f"\nWARNING: {len(unmapped)} speaker name(s) not found in spk_map:")
        for name in sorted(unmapped):
            print(f"    {name}")

    write_word_freq_excel(
        dict(word_freq),
        os.path.join(output_dir, f"{tag}_norm_trans_word_freq.xlsx"),
    )
    print(f"Done in {time.time() - t0:.1f}s")


# ── Mode: extract transcripts ────────────────────────────────────────────────

def run_extract_trans(pdf_dir: str, output_dir: str) -> None:
    """Write one raw transcript per PDF (speaker names as-is) to output_dir."""
    os.makedirs(output_dir, exist_ok=True)

    pdfs = sorted(
        f for f in os.listdir(pdf_dir)
        if f.startswith("transcript_") and f.endswith(".pdf")
    )

    print(f"\nProcessing {len(pdfs)} PDFs -> {output_dir}")
    print("-" * 55)

    for fname in pdfs:
        idx   = int(fname.replace("transcript_", "").replace(".pdf", ""))
        path  = os.path.join(pdf_dir, fname)
        out   = os.path.join(output_dir, f"transcript_{idx:03d}.txt")
        turns = extract_turns(path, {}, set())
        write_transcript([(pdf_name, txt) for _, pdf_name, txt in turns], out)
        print(f"  [{idx:02d}] {fname}  ->  {len(turns)} turns")

    print(f"\nDone. Output: {output_dir}")


# ── Mode: content stats ───────────────────────────────────────────────────────

def run_transcript_content_stats(transcript_dir: str, output_dir: str, tag: str) -> None:
    """Run spaCy content analysis on raw transcripts and write 5-sheet Excel workbook."""
    t0 = time.time()
    print("\n" + "=" * 60)
    print("TRANSCRIPT CONTENT STATISTICS")
    print("=" * 60)
    run_content_stats(transcript_dir, output_dir, tag)
    elapsed  = time.time() - t0
    out_path = os.path.join(output_dir, f"{tag}_transcript_content_analysis.xlsx")
    print(f"\n{'=' * 60}")
    print(f"Done in {elapsed:.1f}s")
    print(f"Output: {out_path}")
    print("=" * 60)


# ── Mode: align and review (V1 vs V3 NSW comparison) ─────────────────────────

# spk_NNNN labels (V1/V3) — and, for --align-and-review, raw-name labels too.
_REVIEW_SPK_RE = re.compile(r"^(spk_\d{4}):\s*$")
_REVIEW_ANY_RE = re.compile(r"^(spk_\d{4}|[A-Z][A-Z0-9 .\-']+):\s*$")

# ── Written-side cleanup for --align-and-review (transcript_original) ─────────
# transcript_original keeps punctuation attached to words ("Please," "Lords.")
# while V3 removes/­spaces it — every such token would otherwise become a spurious
# non-NSW diff ("Other").  We mirror V3's punctuation handling but PROTECT the NSW
# forms (dotted initials A.K. → LSEQ; abbreviations Dr. → Abbreviation).

# Collapse ellipsis / multi-dot runs (2+ dots); single dots are left untouched.
_ELLIPSIS_BEFORE_PUNCT = re.compile(r"\.{2,}\s*(?=[.?!;:,])")
_ELLIPSIS_BEFORE_WORD  = re.compile(r"\.{2,}\s*(?=\S)")
_ELLIPSIS_TRAILING     = re.compile(r"\.{2,}")
# Separators V3 drops -> space.  Kept on purpose: [] (Bracketed NSW), and / ( )
# (slashes/parens carry NSW signal, e.g. dates 25/03/71 and reference forms).
_DROP_SEPARATORS   = re.compile(r"[,;:]")
# Sentence-end ? ! — V3 spaces these away from the word.
_SPACE_BANG_QMARK  = re.compile(r"([^\s?!])([?!]+)")
# Common legal/honorific abbreviations to protect even if absent from the TSV.
_COMMON_ABBR = {
    "mr", "mrs", "ms", "dr", "prof", "hon", "rev", "st", "smt", "sri",
    "vs", "v", "no", "nos", "art", "arts", "sec", "cl", "sch", "para", "paras",
    "vol", "vols", "ch", "pt", "rs", "ltd", "pvt", "co", "ors", "anr", "ano",
    "govt", "dept", "etc", "ie", "eg", "viz", "ibid", "u", "s",
}


def _light_ellipsis_clean(text: str) -> str:
    text = _ELLIPSIS_BEFORE_PUNCT.sub(" ", text)
    text = _ELLIPSIS_BEFORE_WORD.sub(" ", text)
    text = _ELLIPSIS_TRAILING.sub("", text)
    return re.sub(r"\s{2,}", " ", text).strip()


def _space_trailing_period(tok: str, keep_abbr: set) -> str:
    """
    Space a token's trailing sentence-end period ("Lords." -> "Lords ."), but KEEP
    it for NSW forms so they still classify:
      • dotted initials / decimals (internal dot): A.K., U.S.A., 3.14 -> kept
      • single-letter initial: A. -> kept
      • known/common abbreviation: Dr., Rs., No., Art. -> kept
    """
    if not tok.endswith("."):
        return tok
    base = tok[:-1]
    if "." in base:                                   # A.K. / U.S.A. / 3.14
        return tok
    low = base.lower()
    if len(base) == 1 and base.isalpha():             # single initial "A."
        return tok
    if low in keep_abbr or f"{low}." in keep_abbr:    # known abbreviation
        return tok
    return base + " ."                                # sentence-end period


def _clean_original_for_align(text: str, keep_abbr: set) -> str:
    """Written-side preprocessing for --align-and-review (see section header)."""
    text = _light_ellipsis_clean(text)
    text = _DROP_SEPARATORS.sub(" ", text)            # , ; : -> space  (keep / ( ) [])
    text = _SPACE_BANG_QMARK.sub(r"\1 \2", text)      # word? -> word ?
    text = " ".join(_space_trailing_period(t, keep_abbr) for t in text.split())
    return re.sub(r"\s{2,}", " ", text).strip()


def _parse_paired_turns(path: Path, spk_re: re.Pattern = _REVIEW_SPK_RE) -> list[tuple[str, str]]:
    """
    Parse a transcript keeping EVERY turn (including empty bodies).

    nsw_align.parse_transcript drops empty-body turns, which makes the two
    sides' turn counts diverge (a turn that is just <noise> becomes empty in V3).
    Since both sides are written from the same ordered turn list, keeping all turns
    guarantees a 1:1 pairing.  `spk_re` selects the speaker-label format
    (spk_NNNN for V1/V3, or spk_NNNN|RAW-NAME for transcript_original).
    """
    turns:    list[tuple[str, str]] = []
    cur_spk:  str | None            = None
    cur_body: list[str]             = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.rstrip()
        m = spk_re.match(line)
        if m:
            if cur_spk is not None:
                turns.append((cur_spk, " ".join(cur_body).strip()))
            cur_spk, cur_body = m.group(1), []
        elif cur_spk is not None and line.strip():
            cur_body.append(line.strip())
    if cur_spk is not None:
        turns.append((cur_spk, " ".join(cur_body).strip()))
    return turns


# ── NSW review shared helpers (--align-and-review / --align-and-review2) ──────

def _write_review_summary(wb, counts: dict, categories: list[str], fname: str) -> None:
    """
    SUMMARY sheet listing each content-stats NSW category and its count.
    Each category name is an internal hyperlink to that category's sheet.
    """
    ws = wb.create_sheet("SUMMARY", index=0)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2E4057")
    center    = Alignment(horizontal="center")
    link_font = Font(bold=True, color="0563C1", underline="single")

    ws.cell(1, 1, "File:").font = Font(bold=True)
    ws.cell(1, 2, fname)
    for ci, h in enumerate(["Category", "Count"], start=1):
        c = ws.cell(3, ci, h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center

    total = 0
    for ri, cat in enumerate(categories, start=4):
        n = counts.get(cat, 0)
        total += n
        c = ws.cell(ri, 1, cat)
        # internal hyperlink → top of that category's sheet (quote name for spaces)
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{cat}'!A1", display=cat)
        c.font = link_font
        ws.cell(ri, 2, n).alignment = center

    r = 4 + len(categories)
    ws.cell(r, 1, "TOTAL").font = Font(bold=True)
    tc = ws.cell(r, 2, total); tc.font = Font(bold=True); tc.alignment = center


def _run_content_stats_review(written_dir: str, v3_dir: str, file_ids: list[str],
                              output_dir: str, abbrev_tsv: str, *,
                              tag: str, out_prefix: str, spk_re: re.Pattern,
                              written_preprocess, banner: str) -> None:
    """
    Shared driver for --align-and-review (written side = transcript_original) and
    --align-and-review2 (written side = V1).

    Aligns each turn's WRITTEN side against V3, classifies every difference with the
    --content-stats NSW categories (+ 'Other'), and writes <tag>_<out_prefix>-<id>.xlsx
    per file id using utils/nsw_align's alignment and Excel format.
    """
    abbrev_set = load_abbrev_set(Path(abbrev_tsv)) if os.path.exists(abbrev_tsv) else set()
    categories = NSW_CATEGORIES + ["Other"]

    t0 = time.time()
    print("\n" + "=" * 60)
    print(banner)
    print("=" * 60)
    print(f"  Abbreviation entries loaded: {len(abbrev_set)}")

    for raw_id in file_ids:
        fid   = raw_id.strip().zfill(3)
        fname = f"transcript_{fid}.txt"
        w_path  = Path(written_dir) / fname
        v3_path = Path(v3_dir) / fname

        if not w_path.exists() or not v3_path.exists():
            print(f"  [skip] {fname}: missing in written or V3 directory")
            continue

        w_turns  = _parse_paired_turns(w_path, spk_re)
        v3_turns = _parse_paired_turns(v3_path, spk_re)
        if len(w_turns) != len(v3_turns):
            print(f"  [skip] {fname}: turn count mismatch "
                  f"(written={len(w_turns)}, V3={len(v3_turns)})")
            continue

        rows = {cat: [] for cat in categories}
        for idx, ((spk, w_body), (_, v3_body)) in enumerate(zip(w_turns, v3_turns), start=1):
            w_toks  = written_preprocess(w_body).split()
            v3_toks = v3_body.split()
            label   = f"{spk} T{idx}"
            for d in align_turn(w_toks, v3_toks):
                span           = " ".join(d["v1_orig"])
                cat            = classify_nsw_span(span, abbrev_set) or "Other"
                bef_w, aft_w   = get_context(w_toks, d["i1"], d["i2"])
                bef_v3, aft_v3 = get_v3_context(v3_toks, d["j1"], d["j2"])
                rows[cat].append([
                    span, " ".join(d["v3"]),
                    bef_w, aft_w, bef_v3, aft_v3, label,
                ])

        counts   = {cat: len(r) for cat, r in rows.items()}
        out_path = os.path.join(output_dir, f"{tag}_{out_prefix}-{fid}.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _write_review_summary(wb, counts, categories, fname)
        for cat in categories:
            _write_data_sheet(wb, cat, rows[cat])
        wb.save(out_path)
        print(f"  {fname}: {sum(counts.values())} NSW differences  ->  {out_path}")

    print(f"\nDone in {time.time() - t0:.1f}s")


def run_align_and_review(orig_dir: str, v3_dir: str, file_ids: list[str],
                         output_dir: str, abbrev_tsv: str, tag: str) -> None:
    """
    Main NSW review: transcript_original vs V3, classified with the content-stats
    NSW categories → <tag>_nsw-review-<id>.xlsx.

    The written side is transcript_original (case + dots preserved). It is cleaned
    to mirror V3's punctuation handling — ellipsis collapse, drop , ; :, space
    sentence-end . ? ! (keeping / ( ) [] for their NSW signal) — while PROTECTING
    NSW forms (dotted initials A.K. → LSEQ, abbreviations Dr. → Abbreviation),
    unlike V1's clean_v1 which strips them.
    """
    abbrev_set = load_abbrev_set(Path(abbrev_tsv)) if os.path.exists(abbrev_tsv) else set()
    keep_abbr  = abbrev_set | _COMMON_ABBR

    _run_content_stats_review(
        orig_dir, v3_dir, file_ids, output_dir, abbrev_tsv,
        tag=tag,
        out_prefix="nsw-review",
        spk_re=_REVIEW_ANY_RE,                   # transcript_original uses raw-name labels
        written_preprocess=lambda body: _clean_original_for_align(body, keep_abbr),
        banner="ALIGN AND REVIEW  (transcript_original vs V3, content-stats NSW categories)",
    )


def run_align_and_review2(v1_dir: str, v3_dir: str, file_ids: list[str],
                          output_dir: str, abbrev_tsv: str, tag: str) -> None:
    """
    Alternative NSW review: V1 vs V3, using the SAME content-stats NSW categories and
    hyperlinked summary as --align-and-review → <tag>_nsw-review2-<id>.xlsx. Differs only
    in the written side (V1, already normalized by clean_v1, so no extra preprocessing).
    """
    _run_content_stats_review(
        v1_dir, v3_dir, file_ids, output_dir, abbrev_tsv,
        tag=tag,
        out_prefix="nsw-review2",
        spk_re=_REVIEW_SPK_RE,                    # V1 uses spk_NNNN labels
        written_preprocess=lambda t: t,           # V1 is already normalized
        banner="ALIGN AND REVIEW 2  (V1 vs V3, content-stats NSW categories)",
    )


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Analysis pipeline for Supreme Court ASR transcripts"
    )

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--pdf-summary",
        action="store_true",
        help="Extract structural metadata from all PDFs and write pdf-analysis-summary.txt "
             "to analysis-outputs/. Reports page counts, word counts, speaker counts, "
             "font analysis, noise-line counts, unusual speaker names, and character "
             "frequency. Requires --pdf-dir.",
    )
    group.add_argument(
        "--extract-trans",
        action="store_true",
        help="Extract raw transcripts only (speaker names as-is from PDF, no ID resolution). "
             "Writes transcript_NNN.txt files to analysis-outputs/<tag>_transcript_original/. "
             "These feed --content-stats and --dedup-names. Requires --pdf-dir and --tag.",
    )
    group.add_argument(
        "--dedup-names",
        action="store_true",
        help="Speaker name deduplication only. Reads the extracted transcripts in "
             "analysis-outputs/<tag>_transcript_original/, clusters speaker names, and writes "
             "two outputs to analysis-outputs/: <tag>_spk-id-mapping.txt and "
             "<tag>_speaker-name-analysis.xlsx (SHEET1 stats, SHEET2 matrix, SHEET3 dedup "
             "review). Requires --tag (run --extract-trans first).",
    )
    group.add_argument(
        "--get-norm-trans",
        action="store_true",
        help="With --pdf-dir: run the full pipeline from PDFs — speaker analysis, "
             "deduplication, spk_NNNN-mapped transcripts (V1 cleanup + V3 spoken form), "
             "word-frequency, and content-stats. Writes analysis-outputs/ "
             "<tag>_transcript_original/ (raw names), "
             "<tag>_spkid_mapped_transcript_v1/ (V1), "
             "<tag>_spkid_mapped_norm_transcript_v3/ (V3 spoken form), "
             "<tag>_pdf-analysis-summary.txt, <tag>_spk-id-mapping.txt, "
             "<tag>_speaker-name-analysis.xlsx, <tag>_norm_trans_word_freq.xlsx, and "
             "<tag>_transcript_content_analysis.xlsx. "
             "With only --tag (no --pdf-dir): use the already-extracted "
             "<tag>_transcript_original/ to produce the same V1/V3 transcripts, "
             "word-frequency, dedup, and content-stats outputs (everything except "
             "<tag>_pdf-analysis-summary.txt, which needs the PDFs). Requires --tag.",
    )
    group.add_argument(
        "--content-stats",
        action="store_true",
        help="Analyse the extracted transcripts in analysis-outputs/<tag>_transcript_original/ "
             "with spaCy and write a 5-sheet Excel workbook to analysis-outputs/. "
             "Sheets: Overview (speaker count + sentence stats), Category Counts, "
             "Character Frequency, Word Frequency, NSW Frequency. "
             "Requires --tag (run --extract-trans first).",
    )
    group.add_argument(
        "--align-and-review",
        action="store_true",
        help="Main NSW review. For each --file-id, align <tag>_transcript_original/ against "
             "<tag>_spkid_mapped_norm_transcript_v3/ (original case + dots preserved, cleaned "
             "only for ellipsis/punctuation so LSEQ A.K. and Abbreviations Dr. classify "
             "correctly), classify each difference with the --content-stats NSW categories, "
             "and write analysis-outputs/<tag>_nsw-review-<id>.xlsx (one workbook per file). "
             "Requires --tag and --file-id (run --get-norm-trans first).",
    )
    group.add_argument(
        "--align-and-review2",
        action="store_true",
        help="Same as --align-and-review (same content-stats NSW categories and hyperlinked "
             "summary) but the written side is V1 (<tag>_spkid_mapped_transcript_v1/) instead "
             "of transcript_original. Aligns V1 vs <tag>_spkid_mapped_norm_transcript_v3/ and "
             "writes analysis-outputs/<tag>_nsw-review2-<id>.xlsx (one workbook per file). "
             "Requires --tag and --file-id (run --get-norm-trans first).",
    )

    parser.add_argument("--pdf-dir", help="Directory containing transcript_NNN.pdf files "
                                          "(required for --pdf-summary and --extract-trans; "
                                          "optional for --get-norm-trans)")
    parser.add_argument("--file-id", help="Comma-separated transcript ids for --align-and-review[2] "
                                          "(e.g. 001,003). Each id N reviews transcript_N.txt.")
    parser.add_argument("--tag",     help="Short label used for output filenames and the "
                                          "analysis-outputs/<tag>_transcript_original/ directory. "
                                          "Example: 24-file-data")
    args = parser.parse_args()

    # ── --align-and-review[2]: compare a written side vs V3 for given file ids ──
    if args.align_and_review or args.align_and_review2:
        mode = "--align-and-review2" if args.align_and_review2 else "--align-and-review"
        if not args.tag:
            parser.error(f"{mode} requires --tag  (e.g. --tag 24-file-data)")
        if not args.file_id:
            parser.error(f"{mode} requires --file-id  (e.g. --file-id 001,003)")
        v3_dir = os.path.join(OUT_DIR, f"{args.tag}_spkid_mapped_norm_transcript_v3")
        # main review uses transcript_original as the written side; review2 (Flint) uses V1
        if args.align_and_review2:
            written_dir = os.path.join(OUT_DIR, f"{args.tag}_spkid_mapped_transcript_v1")
        else:
            written_dir = os.path.join(OUT_DIR, f"{args.tag}_transcript_original")
        for d in (written_dir, v3_dir):
            if not os.path.isdir(d):
                print(f"ERROR: directory not found: {d}")
                print(f"       Run  --get-norm-trans --pdf-dir <pdfs> --tag {args.tag}  first.")
                sys.exit(1)
        file_ids = [x.strip() for x in args.file_id.split(",") if x.strip()]
        if not file_ids:
            parser.error("--file-id is empty after parsing (e.g. --file-id 001,003)")
        abbrev_tsv = os.path.join(SCRIPT_DIR, "abbrev-expansions.tsv")
        if args.align_and_review2:
            run_align_and_review2(written_dir, v3_dir, file_ids, OUT_DIR, abbrev_tsv, args.tag)
        else:
            run_align_and_review(written_dir, v3_dir, file_ids, OUT_DIR, abbrev_tsv, args.tag)
        return

    # ── Modes that read the extracted transcripts (need --tag, not --pdf-dir) ──
    if args.content_stats or args.dedup_names:
        mode = "--content-stats" if args.content_stats else "--dedup-names"
        if not args.tag:
            parser.error(f"{mode} requires --tag  (e.g. --tag 24-file-data)")
        transcript_dir = os.path.join(OUT_DIR, f"{args.tag}_transcript_original")
        if not os.path.isdir(transcript_dir):
            print(f"ERROR: transcript directory not found: {transcript_dir}")
            print(f"       Run  --extract-trans --pdf-dir <pdfs> --tag {args.tag}  first.")
            sys.exit(1)
        if args.content_stats:
            run_transcript_content_stats(transcript_dir, OUT_DIR, args.tag)
        else:
            run_dedup_names(transcript_dir, OUT_DIR, args.tag)
        return

    # ── --get-norm-trans: full from PDFs (--pdf-dir) OR analysis-only (--tag) ──
    if args.get_norm_trans:
        if not args.tag:
            parser.error("--get-norm-trans requires --tag  (e.g. --tag 24-file-data)")
        os.makedirs(OUT_DIR, exist_ok=True)
        if args.pdf_dir:
            pdf_dir = os.path.normpath(args.pdf_dir)
            if not os.path.isdir(pdf_dir):
                print(f"ERROR: PDF directory not found: {pdf_dir}")
                sys.exit(1)
            AnalysisPipeline(pdf_dir=pdf_dir, output_dir=OUT_DIR, tag=args.tag).run()
            # Also run content-stats on the transcripts just extracted
            transcript_dir = os.path.join(OUT_DIR, f"{args.tag}_transcript_original")
            run_transcript_content_stats(transcript_dir, OUT_DIR, args.tag)
        else:
            # tag-only: analyse the already-extracted transcripts
            transcript_dir = os.path.join(OUT_DIR, f"{args.tag}_transcript_original")
            if not os.path.isdir(transcript_dir):
                print(f"ERROR: transcript directory not found: {transcript_dir}")
                print(f"       Provide --pdf-dir, or run "
                      f"--extract-trans --pdf-dir <pdfs> --tag {args.tag}  first.")
                sys.exit(1)
            deduplicator = run_dedup_names(transcript_dir, OUT_DIR, args.tag)
            write_norm_transcripts(transcript_dir, OUT_DIR, args.tag,
                                   deduplicator.raw_spk_id)
            run_transcript_content_stats(transcript_dir, OUT_DIR, args.tag)
        return

    # ── Remaining modes read the PDFs (need --pdf-dir and --tag) ───────────────
    if not args.pdf_dir:
        parser.error("this mode requires --pdf-dir")
    pdf_dir = os.path.normpath(args.pdf_dir)
    if not os.path.isdir(pdf_dir):
        print(f"ERROR: PDF directory not found: {pdf_dir}")
        sys.exit(1)

    if args.pdf_summary:
        if not args.tag:
            parser.error("--pdf-summary requires --tag  (e.g. --tag 24-file-data)")
        run_pdf_summary(pdf_dir, OUT_DIR, args.tag)
    elif args.extract_trans:
        if not args.tag:
            parser.error("--extract-trans requires --tag  (e.g. --tag 24-file-data)")
        trans_dir = os.path.join(OUT_DIR, f"{args.tag}_transcript_original")
        run_extract_trans(pdf_dir, trans_dir)


if __name__ == "__main__":
    main()
